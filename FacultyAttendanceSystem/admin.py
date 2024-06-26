from django.contrib import admin
from . import models
from .forms import XLSXUploadForm
import pandas as pd # type: ignore
from django.contrib import admin
from .models import Students, StudentClass, Semester
from django.contrib import messages
from django.shortcuts import redirect
from django.urls import path
import csv
from django.contrib import admin
from django.shortcuts import render, redirect
from openpyxl import load_workbook  # type: ignore
import traceback
from datetime import datetime
import openpyxl # type: ignore

@admin.register(models.AdminCredentials)
class AdminCredentialsAdmin(admin.ModelAdmin):
    list_display = ['faculty', 'username']
    search_fields = ['faculty__name', 'faculty__short_name']

@admin.register(models.ClassDuration)
class ClassDurationAdmin(admin.ModelAdmin):
    list_display = ('duration', 'duration_short_name', 'hours', 'minute')

@admin.register(models.Faculty)
class FacultyAdmin(admin.ModelAdmin):
    search_fields = ('name', 'short_name')
    list_display = ('id','name', 'short_name')

@admin.register(models.Subject)
class SubjectAdmin(admin.ModelAdmin):
    list_display = ('name', 'short_name', 'semester')
    search_fields = ('name', 'short_name')

@admin.register(models.Room)
class RoomAdmin(admin.ModelAdmin):
    list_display = ('room_name',)


@admin.register(models.WorkShift)
class WorkshiftAdmin(admin.ModelAdmin):
    list_display = ('faculty', 'formatted_date', 'formatted_punch_in', 'formatted_punch_out')
    search_fields = ('faculty__name', 'faculty__short_name') 
    list_filter = ('faculty__name',) 
    
#custom design: 
    def formatted_date(self, obj):
        """Custom method to display the date in dd/mm/yyyy format."""
        if obj.date:
            return obj.date.strftime('%d/%m/%Y')
        else:
            return '-'
    formatted_date.short_description = 'Date'

    def formatted_punch_in(self, obj):
        """Custom method to display the punch in time with seconds."""
        if obj.punch_in:
            return obj.punch_in.strftime('%H:%M:%S')
        else:
            return '-'
    formatted_punch_in.short_description = 'Punch In'

 
    def formatted_punch_out(self, obj):
      """Custom method to display the punch out time with seconds."""
      if obj.punch_out:
        return obj.punch_out.strftime('%H:%M:%S')
      else:
        return '-'
    formatted_punch_out.short_description = 'Punch Out'

    
@admin.register(models.Timetable)
class TimetableAdmin(admin.ModelAdmin):
    list_display = ('class_type', 'Student_Class','formatted_semester', 'faculty',  'subject', 'room', 'duration', 'start_time', 'end_time', 'formatted_create_date', 'formatted_modified_date')
    list_filter = ('class_type', 'semester', 'faculty', 'subject', 'room', 'create_date', 'modified_date')
    search_fields = ('semester__name', 'faculty__name', 'subject__name', 'room__room_name', 'faculty__short_name')

    def formatted_semester(self, obj):
        if obj.semester:
            parts = obj.semester.name.split('-')
            return f'{parts[0]}-{parts[1]}' if len(parts) >= 2 else parts[0]
        else:
            return '-'

    formatted_semester.short_description = 'Semester'

    def formatted_create_date(self, obj):
        return obj.create_date.strftime("%d/%m/%Y")

    formatted_create_date.short_description = 'Class Definition Create Date'

    def formatted_modified_date(self, obj):
        return obj.modified_date.strftime("%d/%m/%Y")

    formatted_modified_date.short_description = 'Class Definition Modified Date'
 
@admin.register(models.TimeTableRollouts)
class TimeTableRolloutsAdmin(admin.ModelAdmin):
    list_display = ('subject', 'faculty',  'room', 'duration', 'class_status', 'formatted_class_date', 'start_time', 'end_time', 'class_attedance')
    list_filter = ('subject', 'faculty', 'room', 'class_status', 'class_attedance', 'class_date')
    search_fields = ('subject__name', 'faculty__name', 'room__room_name', 'class_date' , 'short_name' ,'faculty__short_name')
    list_editable = ('class_attedance',)

    def formatted_class_date(self, obj):
        return obj.class_date.strftime("%d/%m/%Y")

    formatted_class_date.short_description = 'Class Date'
 
@admin.register(models.Semester)
class SemesterAdmin(admin.ModelAdmin):
    list_display = ('name','term_date','start_date', 'end_date')
    search_fields = ('name',)
    list_filter = ('term_date','start_date', 'end_date')

@admin.register(models.EventScheduler)
class EventSchedulerAdmin(admin.ModelAdmin):
    list_display = ('get_faculty_names', 'date', 'start_time', 'end_time', 'Title', 'Description')
    list_filter = ('date', 'start_time', 'end_time', 'Title')
    search_fields = ('Title', 'Description')
    filter_horizontal = ('faculty',)

    def get_faculty_names(self, obj):
        return ", ".join([faculty.name for faculty in obj.faculty.all()])
    get_faculty_names.short_description = 'Faculty'

@admin.register(models.HolidayScheduler)
class HolidaySchedulerAdmin(admin.ModelAdmin):
    list_display = ('date', 'Title')
    list_filter = ('date', 'Title')
    search_fields = ('Title',)
    
@admin.register(Students)
class StudentsAdmin(admin.ModelAdmin):
    list_display = ('enrollment_no', 'student_name', 'get_students_class_name')
    list_filter = ('Student_Class__Students_class_name',)
    search_fields = ('enrollment_no', 'student_name')
    ordering = ('enrollment_no',)
    change_list_template = "admin/students_changelist.html"

    def get_students_class_name(self, obj):
        return obj.Student_Class.Students_class_name if obj.Student_Class else ''
    get_students_class_name.short_description = 'Class Name'

    def get_urls(self):
        urls = super().get_urls()
        my_urls = [
            path('import-xlsx/', self.import_xlsx),
        ]
        return my_urls + urls

    def import_xlsx(self, request):
        if request.method == "POST":
            form = XLSXUploadForm(request.POST, request.FILES)
            if form.is_valid():
                xlsx_file = request.FILES['xlsx_file']
                try:
                    workbook = openpyxl.load_workbook(xlsx_file)
                    sheet = workbook.active

                    for row in sheet.iter_rows(min_row=2, values_only=True):  # Skip the header row
                        if len(row) < 2:
                            messages.warning(request, f"Skipped row: not enough values (expected 2, got {len(row)})")
                            continue

                        enrollment_no = row[0]
                        student_name = row[1]
                        class_info = row[2] if len(row) > 2 else ''

                        if not enrollment_no or not student_name:
                            messages.error(request, f"Skipped row: Enrollment Number or Student Name missing")
                            continue

                        if not class_info or '-' not in class_info or ' TO ' not in class_info:
                            messages.error(request, f"Skipped row: Invalid class information format")
                            continue

                        # Split class_info into parts
                        class_parts = class_info.split('-')
                        if len(class_parts) < 4:
                            messages.error(request, f"Skipped row: Invalid class information format")
                            continue

                        class_room = '-'.join(class_parts[:-3]).strip()
                        term_name = class_parts[-3].strip()
                        sem_name = class_parts[-2].strip()
                        term_dates = class_parts[-1].strip()

                        start_date_str, end_date_str = term_dates.split(' TO ')
                        start_date = datetime.strptime(start_date_str, '%d-%m-%Y').date()
                        end_date = datetime.strptime(end_date_str, '%d-%m-%Y').date()

                        # Fetch or create Semester object
                        semester, created = Semester.objects.get_or_create(
                            name=sem_name,
                            defaults={'start_date': start_date, 'end_date': end_date}
                        )

                        # Fetch or create StudentClass object
                        student_class, created = StudentClass.objects.get_or_create(
                            Students_class_name=f"{class_room}-{term_name}",
                            semester=semester
                        )

                        # Create Student object
                        Students.objects.create(
                            enrollment_no=enrollment_no,
                            student_name=student_name,
                            Student_Class=student_class
                        )

                    messages.success(request, "XLSX file has been imported successfully.")
                    return redirect("..")

                except Exception as e:
                    messages.error(request, f"Error reading Excel file: {str(e)}")
                    return redirect("..")

        else:
            form = XLSXUploadForm()

        context = {
            'form': form
        }
        return render(request, "admin/xlsx_form.html", context)

    
@admin.register(models.StudentsRollouts)
class StudentsRolloutsAdmin(admin.ModelAdmin):
    list_display = ('student','faculty','subject', 'room', 'start_time', 'end_time', 'formatted_class_date', 'class_status' , 'student_attendance')
    list_filter = ('class_status', 'subject', 'room', 'student')
    search_fields = ('student__student_name', 'subject__name')
    ordering = ('class_date', 'start_time', 'end_time')
    list_editable = ('student_attendance',)


    def formatted_class_date(self, obj):
        return obj.class_date.strftime("%d/%m/%Y")

    fieldsets = (
        (None, {
            'fields': ('student', 'subject', 'room', 'duration', 'class_id', 'class_status')
        }),
        ('Schedule Details', {
            'fields': ('class_date', 'start_time', 'end_time')
        }),
        ('Attendance', {
            'fields': ('student_attendance',)
        }),
        ('Metadata', {
            'fields': ('create_date', 'created_by', 'modified_date', 'modified_by')
        }),
    )
    readonly_fields = ('create_date', 'modified_date')

@admin.register(models.StudentClass)
class StudentClassAdmin(admin.ModelAdmin):
    list_display = ('Students_class_name', 'semester')  # Fields to display in the admin list view
    search_fields = ('Students_class_name',)   